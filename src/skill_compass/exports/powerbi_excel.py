"""Convert the canonical Power BI JSON document to the contract workbook.

This outer export adapter owns Excel typing, named tables, and template style
preservation. It must read only the JSON document for live values and must not
join upstream pipeline files or implement analytical calculations.
"""

from __future__ import annotations

import re
import warnings
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table

from skill_compass.exports.powerbi_contract import PowerBiContractError
from skill_compass.exports.powerbi_json import read_powerbi_json
from skill_compass.schemas.powerbi import (
    PowerBiColumnContract,
    PowerBiExportDocument,
    PowerBiScalar,
)

_EXCEL_TABLE_NAME = re.compile(r"[A-Za-z_][A-Za-z0-9_.]*\Z")
_EXCEL_A1_REFERENCE = re.compile(r"[A-Za-z]{1,3}[1-9][0-9]*\Z")
_EXCEL_R1C1_REFERENCE = re.compile(r"R[1-9][0-9]*C[1-9][0-9]*\Z", re.IGNORECASE)
_INVALID_WORKSHEET_NAME_CHARACTERS = frozenset("[]:*?/\\")
_MAX_EXCEL_COLUMNS = 16_384
_MAX_EXCEL_ROWS = 1_048_576


@dataclass(frozen=True, slots=True)
class PowerBiWorkbookValidation:
    """Summarise a successful post-write Power BI workbook validation."""

    workbook_path: Path
    expected_sheet_count: int
    actual_sheet_count: int
    expected_table_count: int
    actual_table_count: int
    duplicate_table_name_count: int
    invalid_table_range_count: int
    missing_contract_column_count: int
    xml_table_part_count: int


# =============================================================================
# JSON-to-Excel value conversion
# =============================================================================


def _excel_datetime(value: str) -> datetime:
    """Convert an ISO timestamptz value to Excel's UTC-naive datetime type."""
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed


def _excel_value(value: PowerBiScalar, column: PowerBiColumnContract) -> object:
    """Convert one validated JSON scalar to its semantic Excel cell type."""
    if value is None:
        return None
    if column.power_bi_type == "Date":
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    if column.power_bi_type == "Date/Time":
        return _excel_datetime(str(value))
    return value


def _number_format(column: PowerBiColumnContract) -> str:
    """Return an Excel-invariant format for one semantic contract type."""
    if column.power_bi_type == "Date":
        return "yyyy-mm-dd"
    if column.power_bi_type == "Date/Time":
        return "yyyy-mm-dd hh:mm:ss"
    if column.power_bi_type == "Whole Number":
        return "0"
    if column.power_bi_type == "Decimal":
        if "salary" in column.column_name:
            return "#,##0.00"
        return "0.0000"
    if column.power_bi_type == "Text":
        return "@"
    return "General"


# =============================================================================
# Template-preserving table replacement
# =============================================================================


def _validate_excel_table_name(table_name: str) -> None:
    """Reject names outside the conservative Excel table-name safe subset."""
    if (
        len(table_name) > 255
        or _EXCEL_TABLE_NAME.fullmatch(table_name) is None
        or _EXCEL_A1_REFERENCE.fullmatch(table_name) is not None
        or _EXCEL_R1C1_REFERENCE.fullmatch(table_name) is not None
        or table_name.casefold() in {"r", "c"}
    ):
        raise PowerBiContractError(f"invalid Excel table name: {table_name}")


def _validate_worksheet_name(worksheet_name: str) -> None:
    """Reject worksheet names that Excel cannot preserve exactly."""
    if (
        not worksheet_name
        or len(worksheet_name) > 31
        or _INVALID_WORKSHEET_NAME_CHARACTERS.intersection(worksheet_name)
    ):
        raise PowerBiContractError(f"invalid Excel worksheet name: {worksheet_name}")


def _table_range(*, table_name: str, column_count: int, row_count: int) -> str:
    """Build a valid Excel table range, including an insert row when empty.

    Excel represents an empty worksheet table with a header plus a blank insert
    row. The ``insertRow`` table flag distinguishes that row from analytical
    data, so the contract remains logically empty without a synthetic record.
    """
    if column_count < 1:
        raise PowerBiContractError(f"{table_name} must define at least one column")
    if column_count > _MAX_EXCEL_COLUMNS:
        raise PowerBiContractError(f"{table_name} exceeds Excel's column limit")
    final_row = max(2, row_count + 1)
    if final_row > _MAX_EXCEL_ROWS:
        raise PowerBiContractError(f"{table_name} exceeds Excel's row limit")
    return f"A1:{get_column_letter(column_count)}{final_row}"


def _table(worksheet: object, table_name: str) -> Table:
    """Return one expected named table or fail the frozen workbook contract."""
    _validate_excel_table_name(table_name)
    try:
        table = worksheet.tables[table_name]  # type: ignore[attr-defined]
    except KeyError as error:
        raise PowerBiContractError(
            f"reference workbook is missing named table {table_name}"
        ) from error
    if table.name != table_name or table.displayName != table_name:
        raise PowerBiContractError(f"reference workbook renamed table {table_name}")
    return table


def _replace_table(
    *,
    worksheet: object,
    table_name: str,
    columns: tuple[PowerBiColumnContract, ...],
    rows: tuple[dict[str, PowerBiScalar], ...],
) -> None:
    """Replace one table's values while retaining its existing sheet styling."""
    _validate_worksheet_name(worksheet.title)  # type: ignore[attr-defined]
    table_ref = _table_range(
        table_name=table_name,
        column_count=len(columns),
        row_count=len(rows),
    )
    table = _table(worksheet, table_name)
    expected_headers = tuple(column.column_name for column in columns)
    actual_headers = tuple(
        worksheet.cell(row=1, column=index).value  # type: ignore[attr-defined]
        for index in range(1, len(columns) + 1)
    )
    if actual_headers != expected_headers:
        raise PowerBiContractError(f"template headers changed for {table_name}")

    template_styles = []
    for index in range(1, len(columns) + 1):
        cell = worksheet.cell(row=2, column=index)  # type: ignore[attr-defined]
        template_styles.append(
            (
                copy(cell._style),
                copy(cell.alignment),
                copy(cell.protection),
            )
        )
    if worksheet.max_row > 1:  # type: ignore[attr-defined]
        worksheet.delete_rows(2, worksheet.max_row - 1)  # type: ignore[attr-defined]

    for row in rows:
        worksheet.append(  # type: ignore[attr-defined]
            [_excel_value(row[column.column_name], column) for column in columns]
        )

    for row_index in range(2, len(rows) + 2):
        for column_index, column in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)  # type: ignore[attr-defined]
            style, alignment, protection = template_styles[column_index - 1]
            cell._style = copy(style)
            cell.alignment = copy(alignment)
            cell.protection = copy(protection)
            cell.number_format = _number_format(column)

    table.ref = table_ref
    table.totalsRowShown = False
    if rows:
        table.insertRow = None
        table.insertRowShift = None
        if table.autoFilter is not None:
            table.autoFilter.ref = table_ref
    else:
        # This matches the OOXML written by Excel for a genuinely empty table:
        # the ref includes row 2, but insertRow marks it as UI insertion space,
        # not a data record. No placeholder cell values are written.
        table.insertRow = True
        table.insertRowShift = None
        table.autoFilter = AutoFilter(ref=table_ref)
    worksheet.freeze_panes = "A2"  # type: ignore[attr-defined]


def _metadata_column(name: str, type_name: str = "Text") -> PowerBiColumnContract:
    """Create an internal column descriptor for workbook metadata tables."""
    return PowerBiColumnContract(
        view_name="workbook_metadata",
        column_name=name,
        power_bi_type=type_name,
        postgresql_type="text",
        description="Workbook contract metadata.",
        relationship_key=False,
        nullable=False,
    )


def _replace_metadata_tables(workbook: object, document: PowerBiExportDocument) -> None:
    """Populate relationship, dictionary, and live row-count tables from JSON."""
    relationships = tuple(
        relationship.model_dump(mode="python")
        for relationship in document.contract.relationships
    )
    relationship_columns = tuple(
        _metadata_column(name)
        for name in (
            "from_table",
            "from_column",
            "cardinality",
            "to_table",
            "to_column",
            "filter_direction",
            "purpose",
        )
    )
    _replace_table(
        worksheet=workbook["Model_Relationships"],  # type: ignore[index]
        table_name="ModelRelationships",
        columns=relationship_columns,
        rows=relationships,
    )

    dictionary_rows = tuple(
        {
            "excel_table_name": view.view_name,
            "postgresql_view": view.postgresql_view,
            "column_name": column.column_name,
            "power_bi_type": column.power_bi_type,
            "postgresql_type": column.postgresql_type,
            "description": column.description,
            "relationship_key": "Yes" if column.relationship_key else "No",
            "nullable": "Yes" if column.nullable else "No",
        }
        for view in document.contract.views
        for column in view.columns
    )
    dictionary_columns = tuple(
        _metadata_column(name)
        for name in (
            "excel_table_name",
            "postgresql_view",
            "column_name",
            "power_bi_type",
            "postgresql_type",
            "description",
            "relationship_key",
            "nullable",
        )
    )
    _replace_table(
        worksheet=workbook["Data_Dictionary"],  # type: ignore[index]
        table_name="DataDictionary",
        columns=dictionary_columns,
        rows=dictionary_rows,
    )

    summary_rows = tuple(
        {
            "excel_table_name": view.view_name,
            "row_count": len(document.views[view.view_name]),
            "postgresql_target": view.postgresql_view,
            "primary_use": view.primary_use,
        }
        for view in document.contract.views
    )
    summary_columns = (
        _metadata_column("excel_table_name"),
        _metadata_column("row_count", "Whole Number"),
        _metadata_column("postgresql_target"),
        _metadata_column("primary_use"),
    )
    _replace_table(
        worksheet=workbook["Model_Summary"],  # type: ignore[index]
        table_name="ModelSummary",
        columns=summary_columns,
        rows=summary_rows,
    )


def _update_readme(workbook: object, document: PowerBiExportDocument) -> None:
    """Replace synthetic warnings with concise live-export provenance."""
    worksheet = workbook["README"]  # type: ignore[index]
    jobs = document.views["vw_jobs"]
    dates = [row["listing_date"] for row in jobs if row["listing_date"] is not None]
    date_range = f"{min(dates)} to {max(dates)}" if dates else "Not available"
    worksheet["A1"] = "Skill Compass — Live Power BI Contract Export"
    values = {
        5: (
            "Purpose",
            "Load governed live pipeline outputs into the frozen Power BI contract.",
        ),
        6: ("Exported jobs", len(jobs)),
        7: ("Listing-date range", date_range),
        8: ("PostgreSQL target", "Neon PostgreSQL pbi schema"),
        9: (
            "Design principle",
            "Each Excel table and column mirrors the final pbi.vw_* contract.",
        ),
        10: (
            "Data status",
            "Live processed export; unsupported roadmap calculations are intentionally empty.",
        ),
        11: (
            "Power BI migration rule",
            "Keep query names, columns, data types and relationships unchanged; replace only the source step.",
        ),
        12: (
            "DAX caveat",
            "Bridge visuals must use distinct job counts and governed measure definitions.",
        ),
        13: ("Generated", _excel_datetime(document.data_as_of_at)),
    }
    for row_number, (label, value) in values.items():
        worksheet.cell(row=row_number, column=1, value=label)
        worksheet.cell(row=row_number, column=2, value=value)
    worksheet["B13"].number_format = "yyyy-mm-dd hh:mm:ss"


# =============================================================================
# Post-write workbook validation
# =============================================================================


def _load_generated_workbook(path: Path) -> object:
    """Reopen a generated workbook while containing known template warnings."""
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Unknown extension is not supported and will be removed",
                module="openpyxl.worksheet._reader",
            )
            warnings.filterwarnings(
                "ignore",
                message=(
                    "Conditional Formatting extension is not supported and will be removed"
                ),
                module="openpyxl.worksheet._reader",
            )
            return load_workbook(path, data_only=False)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as error:
        raise PowerBiContractError(
            f"generated Power BI workbook could not be reopened: {path}"
        ) from error


def _validate_table_xml_parts(path: Path, expected_count: int) -> int:
    """Require every generated XLSX table part to be present and well-formed XML."""
    try:
        with ZipFile(path) as archive:
            table_parts = sorted(
                name
                for name in archive.namelist()
                if name.startswith("xl/tables/table") and name.endswith(".xml")
            )
            if len(table_parts) != expected_count:
                raise PowerBiContractError(
                    "generated workbook table-part count does not match workbook tables"
                )
            for part_name in table_parts:
                try:
                    root = ElementTree.fromstring(archive.read(part_name))
                except ElementTree.ParseError as error:
                    raise PowerBiContractError(
                        f"generated workbook contains invalid table XML: {part_name}"
                    ) from error
                if root.tag.rsplit("}", maxsplit=1)[-1] != "table":
                    raise PowerBiContractError(
                        f"generated workbook contains an invalid table part: {part_name}"
                    )
    except BadZipFile as error:
        raise PowerBiContractError(
            "generated Power BI workbook is not valid XLSX"
        ) from error
    return len(table_parts)


def validate_powerbi_workbook(
    path: Path, document: PowerBiExportDocument
) -> PowerBiWorkbookValidation:
    """Validate exact sheets, tables, ranges, columns, names, and table XML."""
    workbook = _load_generated_workbook(path)
    expected_sheets = {
        "README",
        "Model_Relationships",
        "Data_Dictionary",
        "Model_Summary",
        *document.contract.view_order,
    }
    if set(workbook.sheetnames) != expected_sheets:  # type: ignore[attr-defined]
        raise PowerBiContractError("generated workbook sheet inventory changed")

    table_names: list[str] = []
    invalid_ranges = 0
    for worksheet in workbook.worksheets:  # type: ignore[attr-defined]
        _validate_worksheet_name(worksheet.title)
        for table_name in worksheet.tables:
            table = worksheet.tables[table_name]
            _validate_excel_table_name(table.name)
            _validate_excel_table_name(table.displayName)
            table_names.append(table.displayName)
            try:
                min_column, min_row, max_column, max_row = range_boundaries(table.ref)
            except ValueError:
                invalid_ranges += 1
                continue
            if (
                min_column < 1
                or min_row < 1
                or max_column < min_column
                or max_row < min_row
                or max_column > _MAX_EXCEL_COLUMNS
                or max_row > _MAX_EXCEL_ROWS
            ):
                invalid_ranges += 1
    duplicate_names = len(table_names) - len({name.casefold() for name in table_names})
    if duplicate_names:
        raise PowerBiContractError("generated workbook contains duplicate table names")
    if invalid_ranges:
        raise PowerBiContractError("generated workbook contains invalid table ranges")

    views_by_name = {view.view_name: view for view in document.contract.views}
    missing_contract_columns = 0
    for view_name in document.contract.view_order:
        worksheet = workbook[view_name]  # type: ignore[index]
        if set(worksheet.tables) != {view_name}:
            raise PowerBiContractError(
                f"generated workbook table inventory changed for {view_name}"
            )
        table = worksheet.tables[view_name]
        if table.name != view_name or table.displayName != view_name:
            raise PowerBiContractError(f"generated workbook renamed table {view_name}")

        expected_columns = tuple(
            column.column_name for column in views_by_name[view_name].columns
        )
        actual_headers = tuple(
            worksheet.cell(row=1, column=index).value
            for index in range(1, len(expected_columns) + 1)
        )
        actual_table_columns = tuple(table.column_names)
        if (
            actual_headers != expected_columns
            or actual_table_columns != expected_columns
        ):
            missing_contract_columns += len(
                set(expected_columns).difference(actual_table_columns)
            )
            raise PowerBiContractError(
                f"generated workbook columns changed for {view_name}"
            )

        min_column, min_row, max_column, max_row = range_boundaries(table.ref)
        expected_max_row = max(2, len(document.views[view_name]) + 1)
        if (min_column, min_row, max_column, max_row) != (
            1,
            1,
            len(expected_columns),
            expected_max_row,
        ):
            raise PowerBiContractError(
                f"generated workbook table range changed for {view_name}"
            )
        if document.views[view_name]:
            if table.insertRow is True:
                raise PowerBiContractError(
                    f"populated table {view_name} is marked as an insert row"
                )
        else:
            if table.insertRow is not True:
                raise PowerBiContractError(
                    f"empty table {view_name} is missing its Excel insert row"
                )
            if table.autoFilter is None or table.autoFilter.ref != table.ref:
                raise PowerBiContractError(
                    f"empty table {view_name} has an invalid auto-filter range"
                )
            if any(
                worksheet.cell(row=2, column=index).value is not None
                for index in range(1, len(expected_columns) + 1)
            ):
                raise PowerBiContractError(
                    f"empty table {view_name} contains a placeholder data value"
                )

    xml_table_part_count = _validate_table_xml_parts(path, len(table_names))
    return PowerBiWorkbookValidation(
        workbook_path=path,
        expected_sheet_count=len(expected_sheets),
        actual_sheet_count=len(workbook.sheetnames),  # type: ignore[attr-defined]
        expected_table_count=len(document.contract.view_order),
        actual_table_count=sum(
            1
            for view_name in document.contract.view_order
            if view_name in workbook[view_name].tables  # type: ignore[index]
        ),
        duplicate_table_name_count=duplicate_names,
        invalid_table_range_count=invalid_ranges,
        missing_contract_column_count=missing_contract_columns,
        xml_table_part_count=xml_table_part_count,
    )


# =============================================================================
# Public Excel conversion boundary
# =============================================================================


def write_powerbi_excel(
    *, json_path: Path, reference_workbook: Path, output_path: Path
) -> PowerBiWorkbookValidation:
    """Convert only the validated JSON values into the reference workbook shape."""
    document = read_powerbi_json(json_path)
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Unknown extension is not supported and will be removed",
                module="openpyxl.worksheet._reader",
            )
            warnings.filterwarnings(
                "ignore",
                message=(
                    "Conditional Formatting extension is not supported and will be removed"
                ),
                module="openpyxl.worksheet._reader",
            )
            workbook = load_workbook(reference_workbook)
    except OSError as error:
        raise PowerBiContractError(
            f"reference workbook could not be opened: {reference_workbook}"
        ) from error

    expected_sheets = {
        "README",
        "Model_Relationships",
        "Data_Dictionary",
        "Model_Summary",
        *document.contract.view_order,
    }
    if set(workbook.sheetnames) != expected_sheets:
        raise PowerBiContractError("reference workbook sheet inventory has changed")

    _replace_metadata_tables(workbook, document)
    views_by_name = {view.view_name: view for view in document.contract.views}
    for view_name in document.contract.view_order:
        _replace_table(
            worksheet=workbook[view_name],
            table_name=view_name,
            columns=views_by_name[view_name].columns,
            rows=document.views[view_name],
        )
    _update_readme(workbook, document)

    generated_at = _excel_datetime(document.data_as_of_at)
    workbook.properties.title = "Skill Compass Live Power BI Contract Export"
    workbook.properties.subject = "Live JSON-to-Excel Power BI contract export"
    workbook.properties.creator = "Skill Compass"
    workbook.properties.lastModifiedBy = "Skill Compass"
    workbook.properties.created = generated_at
    workbook.properties.modified = generated_at
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return validate_powerbi_workbook(output_path, document)
