"""Test Feature 9 JSON-first Power BI export and Excel parity.

These tests use only sanitised fixtures and local files. They must not invoke
external services, write a database, or copy synthetic reference fact values.
"""

from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from skill_compass.cli import main
from skill_compass.exports.powerbi_excel import validate_powerbi_workbook
from skill_compass.exports.powerbi_json import read_powerbi_json
from skill_compass.services.build_analytics import process_analytics
from skill_compass.services.classify_profile_relevance import (
    process_profile_relevance,
)
from skill_compass.services.classify_roles import process_role_classification
from skill_compass.services.classify_seniority import (
    process_seniority_classification,
)
from skill_compass.services.export_powerbi import export_powerbi
from skill_compass.services.extract_requirements import process_cleaned_csv

PROJECT_ROOT = Path(__file__).resolve().parents[2]
FIXTURE = PROJECT_ROOT / "tests/fixtures/cleaned_jobs.csv"
PROFILE = PROJECT_ROOT / "profiles/data_analytics/profile.yaml"
DICTIONARY = PROJECT_ROOT / "profiles/data_analytics/requirements.csv"
ROLE_RULES = PROJECT_ROOT / "profiles/data_analytics/role_rules.yaml"
SENIORITY_RULES = PROJECT_ROOT / "profiles/data_analytics/seniority_rules.yaml"
RELEVANCE_RULES = PROJECT_ROOT / "profiles/data_analytics/relevance_rules.yaml"
REFERENCE_WORKBOOK = (
    PROJECT_ROOT
    / "powerbi/reference/Skill_Compass_Final_Synthetic_PowerBI_Dataset_100_Jobs.xlsx"
)


def _prepare_upstream(input_dir: Path) -> None:
    """Generate the complete local fixture pipeline required by Feature 9."""
    input_dir.mkdir(parents=True)
    cleaned_path = input_dir / "cleaned_jobs.csv"
    shutil.copyfile(FIXTURE, cleaned_path)
    (input_dir / "data_quality_summary.csv").write_text(
        "metric_category,metric_name,metric_value,metric_status,metric_detail\n"
        "run_reconciliation,input_rows,4,info,Sanitised fixture rows.\n"
        "run_reconciliation,duplicate_same_content_rows,0,info,No duplicates.\n"
        "run_reconciliation,reconciliation_pass,true,pass,Fixture reconciled.\n",
        encoding="utf-8",
    )
    process_cleaned_csv(
        input_path=cleaned_path,
        profile_path=PROFILE,
        dictionary_path=DICTIONARY,
        output_dir=input_dir / "skill_extraction",
    )
    process_role_classification(
        input_path=cleaned_path,
        rules_path=ROLE_RULES,
        output_dir=input_dir / "role_classification",
    )
    process_seniority_classification(
        input_path=cleaned_path,
        rules_path=SENIORITY_RULES,
        output_dir=input_dir / "seniority_classification",
    )
    process_profile_relevance(
        input_dir=input_dir,
        rules_path=RELEVANCE_RULES,
        output_dir=input_dir / "profile_relevance",
    )
    process_analytics(
        input_dir=input_dir,
        profile_path=PROFILE,
        dictionary_path=DICTIONARY,
        role_rules_path=ROLE_RULES,
        seniority_rules_path=SENIORITY_RULES,
        output_dir=input_dir / "analytics",
        minimum_sample_size=1,
    )


def _export(input_dir: Path, output_dir: Path) -> Any:
    """Run Feature 9 with every explicit governed configuration path."""
    return export_powerbi(
        input_dir=input_dir,
        output_dir=output_dir,
        reference_workbook=REFERENCE_WORKBOOK,
        profile_path=PROFILE,
        dictionary_path=DICTIONARY,
        role_rules_path=ROLE_RULES,
        seniority_rules_path=SENIORITY_RULES,
    )


def _json_comparable(value: object, power_bi_type: str) -> object:
    """Normalize Excel date values to their canonical JSON representation."""
    if power_bi_type == "Date" and isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def test_export_writes_one_json_then_matching_excel_contract(tmp_path: Path) -> None:
    input_dir = tmp_path / "processed"
    output_dir = tmp_path / "powerbi"
    _prepare_upstream(input_dir)

    result = _export(input_dir, output_dir)
    document = read_powerbi_json(result.json_path)
    workbook = load_workbook(result.excel_path, data_only=True)

    assert result.json_path.is_file()
    assert result.excel_path.is_file()
    assert len(document.views) == 26
    assert sum(len(view.columns) for view in document.contract.views) == 314
    assert document.views["vw_pathway_skill_priorities"] == ()
    assert document.views["vw_roadmap_stages"] == ()
    assert len(document.views["vw_jobs"]) == result.view_row_counts["vw_jobs"]
    assert len(workbook.sheetnames) == 30

    views = {view.view_name: view for view in document.contract.views}
    actual_contract_table_names: set[str] = set()
    for view_name in document.contract.view_order:
        worksheet = workbook[view_name]
        column_contracts = views[view_name].columns
        columns = tuple(column.column_name for column in column_contracts)
        table = worksheet.tables[view_name]
        actual_contract_table_names.add(table.displayName)
        assert tuple(cell.value for cell in worksheet[1]) == columns
        assert table.displayName == view_name
        assert table.ref == (
            f"A1:{get_column_letter(len(columns))}"
            f"{max(2, len(document.views[view_name]) + 1)}"
        )
        assert worksheet.max_row == max(1, len(document.views[view_name]) + 1)
        for excel_row, json_row in zip(
            worksheet.iter_rows(min_row=2, values_only=True),
            document.views[view_name],
            strict=True,
        ):
            expected = tuple(json_row[column] for column in columns)
            actual = tuple(
                _json_comparable(value, column.power_bi_type)
                for value, column in zip(excel_row, column_contracts, strict=True)
            )
            for actual_value, expected_value in zip(actual, expected, strict=True):
                if isinstance(expected_value, str) and "T" in expected_value:
                    assert str(actual_value).startswith(expected_value[:19])
                else:
                    assert actual_value == expected_value

    assert actual_contract_table_names == set(document.contract.view_order)
    validation = validate_powerbi_workbook(result.excel_path, document)
    assert validation.expected_sheet_count == validation.actual_sheet_count == 30
    assert validation.expected_table_count == validation.actual_table_count == 26
    assert validation.duplicate_table_name_count == 0
    assert validation.invalid_table_range_count == 0
    assert validation.missing_contract_column_count == 0
    assert validation.xml_table_part_count == 29

    combined_json = result.json_path.read_text(encoding="utf-8")
    assert "private-" not in combined_json
    assert "fixture-token" not in combined_json
    assert "description_text_clean" not in combined_json
    assert "evidence_snippet" not in combined_json


def test_empty_roadmap_tables_use_excel_native_insert_rows(tmp_path: Path) -> None:
    input_dir = tmp_path / "processed"
    output_dir = tmp_path / "powerbi"
    _prepare_upstream(input_dir)

    result = _export(input_dir, output_dir)
    document = read_powerbi_json(result.json_path)
    workbook = load_workbook(result.excel_path)
    expected_empty_tables = {
        "vw_pathway_skill_priorities": "X",
        "vw_roadmap_stages": "L",
    }

    assert all(document.views[name] == () for name in expected_empty_tables)
    for table_name, final_column in expected_empty_tables.items():
        worksheet = workbook[table_name]
        assert set(worksheet.tables) == {table_name}
        table = worksheet.tables[table_name]
        assert table.name == table.displayName == table_name
        assert table.ref == f"A1:{final_column}2"
        assert table.insertRow is True
        assert table.autoFilter is not None
        assert table.autoFilter.ref == table.ref
        assert all(
            worksheet.cell(row=2, column=column).value is None
            for column in range(1, worksheet.max_column + 1)
        )

    with ZipFile(result.excel_path) as archive:
        table_parts: dict[str, ElementTree.Element] = {}
        for part_name in archive.namelist():
            if part_name.startswith("xl/tables/table") and part_name.endswith(".xml"):
                root = ElementTree.fromstring(archive.read(part_name))
                table_parts[root.attrib["displayName"]] = root
    for table_name, final_column in expected_empty_tables.items():
        assert table_parts[table_name].attrib["ref"] == f"A1:{final_column}2"
        assert table_parts[table_name].attrib["insertRow"] == "1"


def test_repeated_feature_9_export_has_deterministic_json(tmp_path: Path) -> None:
    input_dir = tmp_path / "processed"
    _prepare_upstream(input_dir)

    first = _export(input_dir, tmp_path / "first")
    second = _export(input_dir, tmp_path / "second")

    assert first.json_path.read_bytes() == second.json_path.read_bytes()
    assert first.view_row_counts == second.view_row_counts


def test_export_powerbi_cli_reports_safe_reconciliation(
    tmp_path: Path, capsys: Any
) -> None:
    input_dir = tmp_path / "processed"
    output_dir = tmp_path / "powerbi"
    _prepare_upstream(input_dir)

    exit_code = main(
        [
            "export-powerbi",
            "--input",
            str(input_dir),
            "--output-dir",
            str(output_dir),
            "--reference-workbook",
            str(REFERENCE_WORKBOOK),
            "--profile",
            str(PROFILE),
            "--dictionary",
            str(DICTIONARY),
            "--role-rules",
            str(ROLE_RULES),
            "--seniority-rules",
            str(SENIORITY_RULES),
        ]
    )

    output = capsys.readouterr().out
    assert exit_code == 0
    assert "Contract views: 26" in output
    assert "JSON-to-Excel conversion: PASS" in output
    assert "External API requests: 0" in output
