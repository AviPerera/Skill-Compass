"""Test Power BI Excel table generation at the outer export boundary.

These tests exercise workbook structure only. They must not build analytical
values, change the frozen Power BI contract, or access private source data.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from skill_compass.exports.powerbi_contract import PowerBiContractError
from skill_compass.exports.powerbi_excel import (
    _replace_table,
    _validate_excel_table_name,
)
from skill_compass.schemas.powerbi import PowerBiColumnContract


def _column(view_name: str, column_name: str) -> PowerBiColumnContract:
    """Create one minimal text column for workbook-structure tests."""
    return PowerBiColumnContract(
        view_name=view_name,
        column_name=column_name,
        power_bi_type="Text",
        postgresql_type="text",
        description="Sanitised workbook-structure test column.",
        relationship_key=False,
        nullable=True,
    )


def _template_sheet(workbook: Workbook, table_name: str) -> object:
    """Create one styled two-column template table with a disposable row."""
    worksheet = workbook.create_sheet(table_name)
    worksheet.append(["example_id", "example_name"])
    worksheet.append(["template-id", "Template value"])
    worksheet.add_table(Table(displayName=table_name, ref="A1:B2"))
    return worksheet


def _columns(table_name: str) -> tuple[PowerBiColumnContract, ...]:
    """Return the stable two-column contract used by focused tests."""
    return (
        _column(table_name, "example_id"),
        _column(table_name, "example_name"),
    )


def test_replace_table_writes_populated_contract_table(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    worksheet = _template_sheet(workbook, "vw_example")

    _replace_table(
        worksheet=worksheet,
        table_name="vw_example",
        columns=_columns("vw_example"),
        rows=({"example_id": "1", "example_name": "Example"},),
    )
    output_path = tmp_path / "populated.xlsx"
    workbook.save(output_path)

    reopened = load_workbook(output_path)
    table = reopened["vw_example"].tables["vw_example"]
    assert table.displayName == "vw_example"
    assert table.ref == "A1:B2"
    assert tuple(reopened["vw_example"].values) == (
        ("example_id", "example_name"),
        ("1", "Example"),
    )


def test_replace_table_writes_valid_empty_contract_table(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    worksheet = _template_sheet(workbook, "vw_empty")

    _replace_table(
        worksheet=worksheet,
        table_name="vw_empty",
        columns=_columns("vw_empty"),
        rows=(),
    )
    output_path = tmp_path / "empty.xlsx"
    workbook.save(output_path)

    reopened = load_workbook(output_path)
    table = reopened["vw_empty"].tables["vw_empty"]
    assert table.displayName == "vw_empty"
    assert table.ref == "A1:B2"
    assert table.insertRow is True
    assert tuple(cell.value for cell in reopened["vw_empty"][1]) == (
        "example_id",
        "example_name",
    )
    assert all(
        reopened["vw_empty"].cell(row=2, column=column).value is None
        for column in range(1, 3)
    )


def test_multiple_tables_keep_exact_names_without_suffixes(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    table_names = (
        "vw_jobs",
        "vw_pathway_skill_priorities",
        "vw_roadmap_stages",
    )
    for table_name in table_names:
        worksheet = _template_sheet(workbook, table_name)
        _replace_table(
            worksheet=worksheet,
            table_name=table_name,
            columns=_columns(table_name),
            rows=(),
        )

    output_path = tmp_path / "multiple.xlsx"
    workbook.save(output_path)
    reopened = load_workbook(output_path)

    actual_names = tuple(
        next(iter(reopened[name].tables.values())).displayName for name in table_names
    )
    assert actual_names == table_names
    assert len({name.casefold() for name in actual_names}) == len(table_names)
    assert not any(name[-1].isdigit() for name in actual_names)


def test_replace_table_rejects_zero_column_contract() -> None:
    workbook = Workbook()
    worksheet = workbook.active

    with pytest.raises(PowerBiContractError, match="at least one column"):
        _replace_table(
            worksheet=worksheet,
            table_name="vw_empty",
            columns=(),
            rows=(),
        )


@pytest.mark.parametrize("table_name", ["invalid name", "A1", "R1C1"])
def test_excel_table_name_validation_rejects_unsafe_names(table_name: str) -> None:
    with pytest.raises(PowerBiContractError, match="invalid Excel table name"):
        _validate_excel_table_name(table_name)
